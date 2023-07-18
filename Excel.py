import openpyxl
from tkinter import messagebox
from functools import cache
import Global_Var


class push_excel():
    def __init__(self, path):
        self.path = path
        self.workbook = openpyxl.load_workbook(path)
        try:
            self.sheet = self.workbook["Лист1"]
        except:
            messagebox.showerror("Ошибка", "Неверно выбраны файлы")

    def push_cell(self, data, row_enter, columns_enter, index_data, column_data):
        for column in columns_enter:
            cell = self.sheet.cell(row=row_enter, column=column)
            cell.value = (data.loc[index_data, column_data] / 1000)

    @cache
    def find_index(self, name):
        index = 0
        for cell in self.sheet.iter_cols(min_row=5, max_row=5, values_only=True):
            if cell[0] == name:
                return index
            index += 1

    @cache
    def find_column(self, sub_strs, min_row, max_row):
        index = 1
        indexes = []
        for cell in self.sheet.iter_cols(min_row=min_row, max_row=max_row, values_only=True):
            if cell[0] == None:
                index += 1
                continue
            if sub_strs in str(cell[0]):
                indexes.append(index)
            index += 1
            if len(indexes) == 2:
                break
        return indexes

    @cache
    def find_row_direction_cases(self, Nd_requirements):
        for row in range(6, self.sheet.max_row + 1):
            if Nd_requirements == self.sheet[row][Global_Var.index_Nd].value:
                return row
        return None
