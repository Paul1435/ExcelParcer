from datetime import datetime, timedelta
import pandas as pd
import calendar
import locale
import pymorphy3
import openpyxl
from tkinter import messagebox
from functools import cache


class capconstruction():
    def __init__(self):
        self.pivot_table = None
        self.direction_do = [6]
        self.excluded_prefix = '110203'
        self.exceptions = ['2770', '3034']
        print("init capcon")

    def general_table(self, dfs):
        self.pivot_table = self.otherPivotTable(dfs)
        winter_pivot = self.winterPivotTable(dfs)
        is_mistake_here = False
        for el in self.pivot_table.index:
            if el == "1020-11":
                is_mistake_here = True
        if is_mistake_here:
            self.pivot_table.loc['102-11'] += self.pivot_table.loc['1020-11']
            self.pivot_table = self.pivot_table.drop('1020-11')

        for index in winter_pivot.index:
            self.pivot_table.loc['102-11'] += winter_pivot.loc[index]

    def winterPivotTable(self, dfs):
        filtered_values = set()

        for value in dfs['СПП-элемент']:
            str_val = str(value)
            if str_val.startswith(self.excluded_prefix) and (not str_val[-4:] in self.exceptions):
                filtered_values.add(value)
        filtered_values = list(filtered_values)
        return self.createPivotTable(dfs, filtered_values)

    def otherPivotTable(self, dfs):
        filtered_values = set()

        for value in dfs['СПП-элемент']:
            str_val = str(value)
            if not str_val.startswith(self.excluded_prefix) or (
                    str_val.startswith(self.excluded_prefix) and str_val[-4:] in self.exceptions):
                filtered_values.add(value)

        filtered_values = list(filtered_values)
        return self.createPivotTable(dfs, filtered_values)

    def createPivotTable(self, dfs, filtered_values):
        prePivotTable = dfs.loc[
            (dfs["Напр.Деятельности"].isin(self.direction_do)) & (dfs["СПП-элемент"].isin(filtered_values))]
        pivot_table = pd.pivot_table(prePivotTable, values=['Приход', 'Расход', prePivotTable.columns[6]],
                                     index='КодСлужбыГС',
                                     aggfunc='sum')
        return pivot_table

    @cache
    def find_index(self, sheet, name):
        index = 0
        for cell in sheet.iter_cols(min_row=5, max_row=5, values_only=True):
            if cell[0] == name:
                return index
            index += 1

    @cache
    def find_column(self, sheet, sub_strs):
        index = 1
        indexes = []
        for cell in sheet.iter_cols(min_row=5, max_row=5, values_only=True):
            if cell[0] == None:
                index += 1
                continue
            if sub_strs in str(cell[0]):
                indexes.append(index)
            index += 1
            if len(indexes) == 2:
                break
        return indexes

    def current_Time(self, date_obj):
        locale.setlocale(locale.LC_ALL, 'ru_RU.UTF-8')
        parsed_month = pymorphy3.MorphAnalyzer().parse(calendar.month_name[date_obj.month])[0]
        month_name_decl = parsed_month.inflect({'sing', 'gent'}).word
        return f"{date_obj.day} {month_name_decl} {date_obj.year}"

    def parseTimePivotTable(self):
        locale.setlocale(locale.LC_ALL, 'ru_RU.UTF-8')
        Data = str(self.pivot_table.columns[2]).split(' ')[1]
        date_obj = datetime.strptime(Data, "%d.%m.%Y")
        self.cur_time = self.current_Time(date_obj)
        next_month = (date_obj + timedelta(days=int(Data[:2]))).replace(day=1)
        return self.current_Time(next_month)

    @cache
    def find_row(self, sheet, Nd_requirements, GS_requirements, dta_requirements, fact_requirements):
        for row in range(6, sheet.max_row + 1):
            if Nd_requirements == sheet[row][self.index_Nd].value and GS_requirements == sheet[row][
                self.index_Service_Gov].value and dta_requirements == sheet[row][
                self.index_direction_to_action].value and fact_requirements == sheet[row][self.index_None].value:
                return row
        return None

    def find_const_index(self, sheet):
        self.index_Nd = self.find_index(sheet, "НД")
        self.index_Service_Gov = self.find_index(sheet, "Служба ГС")
        self.index_direction_to_action = self.find_index(sheet, "Направление деятельности")
        self.index_None = self.find_index(sheet, None)

    def input_cell(self, sheet, row, columns, index, pivot_column):
        for column in columns:
            cell = sheet.cell(row=row, column=column)
            cell.value = (self.pivot_table.loc[index, pivot_column] / 1000)

    @cache
    def add_value_excel(self, need_period, path):
        workbook = openpyxl.load_workbook(path)
        try:
            sheet = workbook["Лист1"]
        except:
            messagebox.showerror("Ошибка", "Неверно выбраны файлы")
        self.find_const_index(sheet)

        for index in self.pivot_table.index:
            row = self.find_row(sheet, "КС", index, "текущий запас", "факт")
            columns_reserve = self.find_column(sheet, "Запасы на " + str(need_period))
            columns_profit = self.find_column(sheet, "Приход " + str(self.cur_time)[3:])
            columns_lost = []
            if len(columns_profit) != 0:
                columns_lost = self.find_column(sheet, "Расход " + str(self.cur_time)[3:])
            self.input_cell(sheet, row, columns_reserve, index, self.pivot_table.columns[2])
            self.input_cell(sheet, row, columns_profit, index, "Приход")
            self.input_cell(sheet, row, columns_lost, index, "Расход")
        try:
            workbook.save(path)
        except:
            messagebox.showerror("Ошибка", "Нет доступа к файлу " + path + " вероятно он открыт.")
        print("Successful enter KC")

    def automatic(self, obj, template_obj):
        self.general_table(obj)
        need_period = self.parseTimePivotTable()
        self.add_value_excel(need_period, template_obj)
